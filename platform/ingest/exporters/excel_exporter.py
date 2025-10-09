"""
Excel Exporter

Exports financial data to Excel files using YAML cell mappings.
Supports round-trip: database → export Excel with same structure as import.
"""

import yaml
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export financial data to Excel files using YAML cell maps.
    
    Mirrors the structure of ExcelImporter for round-trip compatibility.
    """
    
    def __init__(self, cell_map_path: str):
        """
        Initialize exporter with YAML cell map.
        
        Args:
            cell_map_path: Path to YAML cell map file
        """
        with open(cell_map_path, 'r') as f:
            self.cell_map = yaml.safe_load(f)
        
        self.workbook = None
    
    def create_workbook(self) -> None:
        """Create a new Excel workbook."""
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        logger.info("Created new workbook")
    
    def apply_header_style(self, sheet, row_num: int, col_count: int) -> None:
        """Apply header styling to a row."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, col_count + 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def apply_data_style(self, sheet, start_row: int, end_row: int, col_count: int) -> None:
        """Apply data styling to rows."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def export_consensus_estimates(self, data: List[Dict[str, Any]]) -> None:
        """
        Export consensus estimates to Excel.
        
        Args:
            data: List of consensus estimate records
        """
        if 'consensus_estimates' not in self.cell_map:
            raise ValueError("No consensus_estimates mapping in cell map")
        
        mapping = self.cell_map['consensus_estimates']
        sheet_name = mapping['sheet']
        columns = mapping['columns']
        
        # Create sheet
        sheet = self.workbook.create_sheet(sheet_name)
        
        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = col_def['name'].replace('_', ' ').title()
        
        # Apply header style
        self.apply_header_style(sheet, 1, len(columns))
        
        # Write data
        for row_idx, record in enumerate(data, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = record.get(col_def['name'])
                cell.value = value
                
                # Apply number format for floats
                if col_def.get('type') == 'float' and value is not None:
                    cell.number_format = '#,##0.00'
        
        # Apply data style
        if data:
            self.apply_data_style(sheet, 2, len(data) + 1, len(columns))
        
        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        
        logger.info(f"Exported {len(data)} consensus estimates")
    
    def export_price_targets(self, data: List[Dict[str, Any]]) -> None:
        """
        Export price targets to Excel.
        
        Args:
            data: List of price target records
        """
        if 'price_targets' not in self.cell_map:
            raise ValueError("No price_targets mapping in cell map")
        
        mapping = self.cell_map['price_targets']
        sheet_name = mapping['sheet']
        columns = mapping['columns']
        
        # Create sheet
        sheet = self.workbook.create_sheet(sheet_name)
        
        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = col_def['name'].replace('_', ' ').title()
        
        # Apply header style
        self.apply_header_style(sheet, 1, len(columns))
        
        # Write data
        for row_idx, record in enumerate(data, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = record.get(col_def['name'])
                cell.value = value
                
                # Apply formats
                if col_def.get('type') == 'float' and value is not None:
                    cell.number_format = '$#,##0.00'
                elif col_def.get('type') == 'date' and value is not None:
                    cell.number_format = 'yyyy-mm-dd'
        
        # Apply data style
        if data:
            self.apply_data_style(sheet, 2, len(data) + 1, len(columns))
        
        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 18
        
        logger.info(f"Exported {len(data)} price targets")
    
    def export_loe_events(self, data: List[Dict[str, Any]]) -> None:
        """
        Export LoE events to Excel.
        
        Args:
            data: List of LoE event records
        """
        if 'loe_events' not in self.cell_map:
            raise ValueError("No loe_events mapping in cell map")
        
        mapping = self.cell_map['loe_events']
        sheet_name = mapping['sheet']
        columns = mapping['columns']
        
        # Create sheet
        sheet = self.workbook.create_sheet(sheet_name)
        
        # Write headers
        for idx, col_def in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = col_def['name'].replace('_', ' ').title()
        
        # Apply header style
        self.apply_header_style(sheet, 1, len(columns))
        
        # Write data
        for row_idx, record in enumerate(data, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = record.get(col_def['name'])
                cell.value = value
                
                # Apply formats
                if col_def.get('type') == 'float' and value is not None:
                    cell.number_format = '#,##0.00'
                elif col_def.get('type') == 'date' and value is not None:
                    cell.number_format = 'yyyy-mm-dd'
        
        # Apply data style
        if data:
            self.apply_data_style(sheet, 2, len(data) + 1, len(columns))
        
        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        
        logger.info(f"Exported {len(data)} LoE events")
    
    def export_all(self, data: Dict[str, List[Dict[str, Any]]]) -> None:
        """
        Export all data types to Excel.
        
        Args:
            data: Dictionary with all data to export
        """
        if not self.workbook:
            self.create_workbook()
        
        if 'consensus_estimates' in data and 'consensus_estimates' in self.cell_map:
            self.export_consensus_estimates(data['consensus_estimates'])
        
        if 'price_targets' in data and 'price_targets' in self.cell_map:
            self.export_price_targets(data['price_targets'])
        
        if 'loe_events' in data and 'loe_events' in self.cell_map:
            self.export_loe_events(data['loe_events'])
    
    def save(self, excel_path: str) -> None:
        """
        Save workbook to file.
        
        Args:
            excel_path: Output Excel file path
        """
        if not self.workbook:
            raise ValueError("No workbook to save")
        
        self.workbook.save(excel_path)
        logger.info(f"Saved workbook to: {excel_path}")
