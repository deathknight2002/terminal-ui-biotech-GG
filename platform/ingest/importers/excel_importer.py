"""
Excel Model Importer

Imports financial models from Excel files using YAML cell mappings.
Supports round-trip: import Excel → database → export Excel with same structure.
"""

import yaml
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


class ExcelImporter:
    """
    Import financial data from Excel files using YAML cell maps.
    
    The YAML cell map defines:
    - Sheet names
    - Cell ranges for each data type
    - Data types and validation
    - Formatting rules
    """
    
    def __init__(self, cell_map_path: str):
        """
        Initialize importer with YAML cell map.
        
        Args:
            cell_map_path: Path to YAML cell map file
        """
        with open(cell_map_path, 'r') as f:
            self.cell_map = yaml.safe_load(f)
        
        self.workbook = None
    
    def load_workbook(self, excel_path: str) -> None:
        """Load Excel workbook."""
        self.workbook = load_workbook(excel_path, data_only=False)
        logger.info(f"Loaded workbook: {excel_path}")
    
    def extract_value(self, sheet_name: str, cell_ref: str) -> Any:
        """
        Extract value from specific cell.
        
        Args:
            sheet_name: Worksheet name
            cell_ref: Cell reference (e.g., 'A1', 'B5')
        
        Returns:
            Cell value
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded")
        
        sheet = self.workbook[sheet_name]
        return sheet[cell_ref].value
    
    def extract_range(self, sheet_name: str, range_ref: str) -> List[List[Any]]:
        """
        Extract values from cell range.
        
        Args:
            sheet_name: Worksheet name
            range_ref: Range reference (e.g., 'A1:C10')
        
        Returns:
            2D list of values
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded")
        
        sheet = self.workbook[sheet_name]
        range_cells = sheet[range_ref]
        
        values = []
        for row in range_cells:
            row_values = [cell.value for cell in row]
            values.append(row_values)
        
        return values
    
    def extract_table(self, sheet_name: str, table_def: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Extract table data into list of dictionaries.
        
        Args:
            sheet_name: Worksheet name
            table_def: Table definition from cell map
        
        Returns:
            List of row dictionaries
        """
        range_ref = table_def['range']
        columns = table_def['columns']
        
        raw_data = self.extract_range(sheet_name, range_ref)
        
        # Skip header row if specified
        if table_def.get('has_header', True):
            raw_data = raw_data[1:]
        
        # Convert to dictionaries
        table_data = []
        for row in raw_data:
            if all(v is None for v in row):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, col_def in enumerate(columns):
                if i < len(row):
                    value = row[i]
                    # Apply type conversion
                    if col_def.get('type') == 'float':
                        value = float(value) if value is not None else None
                    elif col_def.get('type') == 'int':
                        value = int(value) if value is not None else None
                    elif col_def.get('type') == 'date':
                        # Handle date conversion
                        if isinstance(value, datetime):
                            value = value.date()
                    
                    row_dict[col_def['name']] = value
            
            table_data.append(row_dict)
        
        return table_data
    
    def import_consensus_estimates(self) -> List[Dict[str, Any]]:
        """
        Import consensus estimates from Excel.
        
        Returns:
            List of consensus estimate records
        """
        if 'consensus_estimates' not in self.cell_map:
            raise ValueError("No consensus_estimates mapping in cell map")
        
        mapping = self.cell_map['consensus_estimates']
        sheet_name = mapping['sheet']
        
        estimates = self.extract_table(sheet_name, mapping)
        
        logger.info(f"Imported {len(estimates)} consensus estimates")
        return estimates
    
    def import_price_targets(self) -> List[Dict[str, Any]]:
        """
        Import price targets from Excel.
        
        Returns:
            List of price target records
        """
        if 'price_targets' not in self.cell_map:
            raise ValueError("No price_targets mapping in cell map")
        
        mapping = self.cell_map['price_targets']
        sheet_name = mapping['sheet']
        
        targets = self.extract_table(sheet_name, mapping)
        
        logger.info(f"Imported {len(targets)} price targets")
        return targets
    
    def import_loe_events(self) -> List[Dict[str, Any]]:
        """
        Import LoE events from Excel.
        
        Returns:
            List of LoE event records
        """
        if 'loe_events' not in self.cell_map:
            raise ValueError("No loe_events mapping in cell map")
        
        mapping = self.cell_map['loe_events']
        sheet_name = mapping['sheet']
        
        events = self.extract_table(sheet_name, mapping)
        
        logger.info(f"Imported {len(events)} LoE events")
        return events
    
    def import_all(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Import all data types from Excel.
        
        Returns:
            Dictionary with all imported data
        """
        results = {}
        
        if 'consensus_estimates' in self.cell_map:
            results['consensus_estimates'] = self.import_consensus_estimates()
        
        if 'price_targets' in self.cell_map:
            results['price_targets'] = self.import_price_targets()
        
        if 'loe_events' in self.cell_map:
            results['loe_events'] = self.import_loe_events()
        
        return results


# Example cell map YAML structure
EXAMPLE_CELL_MAP = """
# Cell map for financial model Excel import
version: "1.0"

consensus_estimates:
  sheet: "Consensus Estimates"
  range: "A2:F100"
  has_header: true
  columns:
    - name: ticker
      type: string
    - name: metric
      type: string
    - name: period
      type: string
    - name: value
      type: float
    - name: source
      type: string
    - name: currency
      type: string

price_targets:
  sheet: "Price Targets"
  range: "A2:E50"
  has_header: true
  columns:
    - name: source
      type: string
    - name: date
      type: date
    - name: price_target
      type: float
    - name: rationale
      type: string
    - name: currency
      type: string

loe_events:
  sheet: "LoE Timeline"
  range: "A2:H30"
  has_header: true
  columns:
    - name: asset_id
      type: string
    - name: asset_name
      type: string
    - name: region
      type: string
    - name: expiry_date
      type: date
    - name: exclusivity_type
      type: string
    - name: peak_revenue
      type: float
    - name: year_1_erosion
      type: float
    - name: year_2_erosion
      type: float
"""
